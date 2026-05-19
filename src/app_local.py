"""
app_local.py  –  Run locally.  Reads & writes the Excel file from STOCK_EXCEL_PATH.

Usage:
    streamlit run app_local.py

Required env / .env:
    KITE_API_KEY=...
    STOCK_EXCEL_PATH=...path/to/Stocks_2026_Latest.xlsx
    STOCK_SYMBOL_OVERRIDES=...path/to/symbol_overrides.json   (optional, defaults to ./symbol_overrides.json)
"""

from __future__ import annotations

import time
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.views import Selection

# ── project imports (flat, no package prefix) ────────────────────────────────
from config import load_settings
from kite_service import KiteService, DailyClose
from symbols import load_overrides

# ─────────────────────────────────────────────────────────────────────────────
# Style constants  (same palette as excel_updater.py)
# ─────────────────────────────────────────────────────────────────────────────
POSITIVE_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
NEGATIVE_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
NEUTRAL_FILL  = PatternFill(fill_type="solid", fgColor="FFEB9C")
POSITIVE_FONT = Font(color="006100")
NEGATIVE_FONT = Font(color="9C0006")
NEUTRAL_FONT  = Font(color="9C6500")

DATE_COL_MIN_WIDTH = 20.0
SHEET_ZOOM = 150            # open the sheet at 150% instead of 100%
VISIBLE_CONTEXT_COLS = 4    # how many older date columns to keep on screen
                            # alongside the latest one when the file opens


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _parse_date(value: object) -> date | None:
    """Robustly parse a cell value to a date."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 30000:
        try:
            return from_excel(value).date()
        except Exception:
            return None
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y",
                    "%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _get_last_date(ws) -> date | None:
    """Find the rightmost date header in row 1."""
    last = None
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            parsed = _parse_date(cell.value)
            if parsed:
                last = parsed
    return last


def _weekdays_between(start: date, end: date) -> list[date]:
    """Return all Mon-Fri dates from start to end inclusive."""
    out, cur = [], start
    while cur <= end:
        if cur.weekday() < 5:
            out.append(cur)
        cur += timedelta(days=1)
    return out


def _ensure_min_width(ws, col_idx: int) -> None:
    letter = get_column_letter(col_idx)
    current = ws.column_dimensions[letter].width
    if current is None or current < DATE_COL_MIN_WIDTH:
        ws.column_dimensions[letter].width = DATE_COL_MIN_WIDTH


def _find_or_make_date_col(ws, target: date) -> int:
    """Return column index for target date, creating header if needed."""
    last_used = 1
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v not in (None, ""):
            last_used = col
        if _parse_date(v) == target:
            _ensure_min_width(ws, col)
            return col
    new_col = last_used + 1
    hdr = ws.cell(row=1, column=new_col)
    prev = ws.cell(row=1, column=last_used)
    if prev.has_style:
        hdr.font       = copy(prev.font)
        hdr.fill       = copy(prev.fill)
        hdr.border     = copy(prev.border)
        hdr.alignment  = copy(prev.alignment)
        hdr.protection = copy(prev.protection)
    hdr.value         = target
    hdr.number_format = "dd-mm-yy"
    _ensure_min_width(ws, new_col)
    return new_col


def _copy_style(src, tgt) -> None:
    if src.has_style:
        tgt.font       = copy(src.font)
        tgt.fill       = copy(src.fill)
        tgt.border     = copy(src.border)
        tgt.alignment  = copy(src.alignment)
        tgt.number_format = src.number_format
        tgt.protection = copy(src.protection)


def _last_used_col(ws) -> int:
    """Index of the rightmost column that has a header in row 1."""
    last = 1
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value not in (None, ""):
            last = col
    return last


def _configure_view(ws, last_col: int) -> None:
    """Freeze the label column + header row, zoom in, and scroll the view so
    the most recent dates are visible the moment the workbook is opened."""
    ws.freeze_panes = "B2"                       # pin column A + row 1
    ws.sheet_view.zoomScale       = SHEET_ZOOM
    ws.sheet_view.zoomScaleNormal = SHEET_ZOOM

    if last_col > 1 and ws.sheet_view.pane is not None:
        first_visible = max(2, last_col - VISIBLE_CONTEXT_COLS)
        last_cell = f"{get_column_letter(last_col)}2"
        ws.sheet_view.pane.topLeftCell = f"{get_column_letter(first_visible)}2"
        ws.sheet_view.selection = [
            Selection(pane="bottomRight", activeCell=last_cell, sqref=last_cell)
        ]


def _write_price(ws, row: int, col: int, close: DailyClose) -> None:
    """Write a coloured price cell."""
    pct  = close.change_pct
    sign = "+" if pct >= 0 else ""
    cell = ws.cell(row=row, column=col)

    # copy style from nearest left neighbour with data
    for prev_col in range(col - 1, 1, -1):
        prev = ws.cell(row=row, column=prev_col)
        if prev.value not in (None, ""):
            _copy_style(prev, cell)
            break

    cell.value     = f"{close.close:.2f} [{sign}{pct:.2f}%]"
    cell.alignment = Alignment(horizontal="right")

    if pct > 0:
        cell.fill = copy(POSITIVE_FILL)
        cell.font = Font(color="006100",
                         name=cell.font.name, sz=cell.font.sz,
                         b=cell.font.b, i=cell.font.i)
    elif pct < 0:
        cell.fill = copy(NEGATIVE_FILL)
        cell.font = Font(color="9C0006",
                         name=cell.font.name, sz=cell.font.sz,
                         b=cell.font.b, i=cell.font.i)
    else:
        cell.fill = copy(NEUTRAL_FILL)
        cell.font = Font(color="9C6500",
                         name=cell.font.name, sz=cell.font.sz,
                         b=cell.font.b, i=cell.font.i)


def _build_instrument_map(kite: KiteService) -> dict[str, int]:
    """Build a combined NSE + BSE symbol → token map."""
    result = {}
    for exchange in ("NSE", "BSE"):
        for item in kite.client.instruments(exchange):
            sym = str(item.get("tradingsymbol", "")).upper()
            tok = item.get("instrument_token")
            if sym and tok:
                result.setdefault(sym, int(tok))
    return result


def _resolve_token(sym: str, inst_map: dict[str, int]) -> int | None:
    """Look up token, trying plain symbol then -EQ / -BE suffixes."""
    s = sym.upper()
    return inst_map.get(s) or inst_map.get(f"{s}-EQ") or inst_map.get(f"{s}-BE")


# ─────────────────────────────────────────────────────────────────────────────
# Core fill engine  –  O(stocks) API calls for any date range
# ─────────────────────────────────────────────────────────────────────────────

def fill_missing_dates(
    wb,
    sheet_name: str,
    dates_to_fill: list[date],
    kite: KiteService,
    overrides: dict[str, str],
    progress_cb=None,
) -> dict:
    """
    For every stock in the sheet, fetch the entire missing date range in ONE
    API call per stock (not one call per stock×date).  Holiday detection is
    automatic: if a weekday has no candle data, it is simply skipped.

    Returns a summary dict:
        {
          date: {"status": "trading"|"holiday", "updated": int, "skipped": int},
          ...
          "stocks_failed": [(label, reason), ...]
        }
    """
    ws = wb[sheet_name]

    # Collect stock rows (row_index, label)
    stock_rows: list[tuple[int, str]] = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and val.strip():
            stock_rows.append((row, val.strip()))

    if not dates_to_fill or not stock_rows:
        return {}

    start_dt = dates_to_fill[0]
    end_dt   = dates_to_fill[-1]

    # Build instrument map (NSE + BSE)
    inst_map = _build_instrument_map(kite)

    # ── Pass 1: fetch data for every stock into memory ───────────────────────
    # We do NOT create date columns yet — only dates that turn out to have
    # actual data will get a column. This way, holidays / "today before market
    # close" do not create empty header columns that the next run would
    # mistakenly skip.
    fetched: dict[int, dict[date, DailyClose]] = {row_idx: {} for row_idx, _ in stock_rows}
    stocks_failed: list[tuple[str, str]] = []

    total = len(stock_rows)
    for idx, (row_idx, label) in enumerate(stock_rows):
        if progress_cb:
            progress_cb(idx / total * 0.9, f"Fetching {label} ({idx+1}/{total})…")

        symbol = overrides.get(label, "").upper()
        if not symbol:
            stocks_failed.append((label, "no symbol override defined"))
            continue

        token = _resolve_token(symbol, inst_map)
        if not token:
            stocks_failed.append((label, f"symbol '{symbol}' not found in NSE/BSE instruments"))
            continue

        try:
            records = kite.client.historical_data(
                instrument_token=token,
                from_date=datetime.combine(start_dt - timedelta(days=1),
                                           datetime.min.time()),
                to_date=datetime.combine(end_dt, datetime.max.time()),
                interval="day",
            )
            time.sleep(kite.settings.request_delay_seconds)

            by_date: dict[date, dict] = {}
            for rec in records:
                d = rec["date"].date() if isinstance(rec["date"], datetime) else rec["date"]
                by_date[d] = rec

            for target_date in dates_to_fill:
                rec = by_date.get(target_date)
                if rec is not None and float(rec["close"]) > 0:
                    fetched[row_idx][target_date] = DailyClose(
                        symbol=symbol,
                        trading_date=target_date,
                        close=float(rec["close"]),
                        open_price=float(rec["open"]),
                    )

        except Exception as exc:
            stocks_failed.append((label, f"API error: {exc}"))

    # ── Pass 2: create columns only for dates that received data, in order ───
    if progress_cb:
        progress_cb(0.95, "Writing to workbook…")

    dates_with_data = sorted({d for row_data in fetched.values() for d in row_data})
    date_to_col: dict[date, int] = {d: _find_or_make_date_col(ws, d) for d in dates_with_data}

    summary: dict = {d: {"status": "holiday", "updated": 0, "skipped": 0}
                     for d in dates_to_fill}
    for d in dates_with_data:
        summary[d]["status"] = "trading"

    for row_idx, row_data in fetched.items():
        for target_date, close_obj in row_data.items():
            _write_price(ws, row_idx, date_to_col[target_date], close_obj)
            summary[target_date]["updated"] += 1

    for d in dates_to_fill:
        summary[d]["skipped"] = len(stock_rows) - summary[d]["updated"]

    _configure_view(ws, _last_used_col(ws))

    summary["stocks_failed"] = stocks_failed
    if progress_cb:
        progress_cb(1.0, "Done.")
    return summary


# ─────────────────────────────────────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    st.set_page_config(page_title="Stock Excel Manager – Local", page_icon="📊")
    st.title("📊 Stock Report Updater – Local")
    st.caption("Reads and writes directly to your local Excel file.")

    # ── Load settings ─────────────────────────────────────────────────────────
    try:
        settings = load_settings()
    except RuntimeError as e:
        st.error(str(e))
        st.stop()

    # ── Auto-handle Zerodha OAuth callback (?request_token=… in URL) ──────────
    qp = st.query_params
    if "request_token" in qp and "kite_access_token" not in st.session_state:
        rtok = qp["request_token"]
        if isinstance(rtok, list):
            rtok = rtok[0]
        if not settings.kite_api_secret:
            st.error("KITE_API_SECRET missing in .env — cannot auto-exchange request token.")
            st.query_params.clear()
        else:
            try:
                from kiteconnect import KiteConnect
                kc   = KiteConnect(api_key=settings.kite_api_key)
                data = kc.generate_session(rtok, api_secret=settings.kite_api_secret)
                st.session_state["kite_access_token"] = data["access_token"]
                st.session_state["kite_user_name"]    = data.get("user_name", "")
                st.query_params.clear()
                st.rerun()
            except Exception as e:
                st.error(f"Auto-login failed: {e}")
                st.query_params.clear()

    # ── Sidebar ───────────────────────────────────────────────────────────────
    st.sidebar.header("⚙️ Configuration")

    if "kite_access_token" in st.session_state:
        name = st.session_state.get("kite_user_name", "")
        st.sidebar.success(f"✓ Logged in{(' as ' + name) if name else ''}")
        if st.sidebar.button("🚪 Log out", use_container_width=True):
            st.session_state.pop("kite_access_token", None)
            st.session_state.pop("kite_user_name", None)
            st.rerun()
        kite_token = st.session_state["kite_access_token"]
    else:
        login_url = (
            f"https://kite.zerodha.com/connect/login?"
            f"api_key={settings.kite_api_key}&v=3"
        )
        st.sidebar.link_button("🔐 Login with Zerodha", login_url, use_container_width=True)
        st.sidebar.caption(
            "On https://developers.kite.trade/apps set the Redirect URL to "
            "`http://localhost:8501/` for this to work."
        )
        with st.sidebar.expander("Or paste access token manually"):
            manual = st.text_input("Access Token", type="password", key="manual_tok")
            if st.button("Use this token", key="manual_use") and manual.strip():
                st.session_state["kite_access_token"] = manual.strip()
                st.rerun()
        kite_token = ""

    workbook_path = settings.workbook_path
    sheet_name    = settings.sheet_name

    api_key_preview = (settings.kite_api_key[:4] + "…" + settings.kite_api_key[-2:]) \
        if len(settings.kite_api_key) > 6 else "(empty!)"
    st.info(
        f"**Workbook:** `{workbook_path}`  \n"
        f"**Sheet:** `{sheet_name}`  \n"
        f"**Kite API key (from .env):** `{api_key_preview}`"
    )

    if not workbook_path.exists():
        st.error(f"File not found: `{workbook_path}`\n\nCheck `STOCK_EXCEL_PATH` in your `.env`.")
        st.stop()

    # ── Preview: detect missing dates ─────────────────────────────────────────
    try:
        wb_preview = load_workbook(workbook_path, read_only=True, data_only=True)
        ws_preview = wb_preview[sheet_name]
        last_date  = _get_last_date(ws_preview)
        wb_preview.close()
    except Exception as e:
        st.error(f"Could not open workbook: {e}")
        st.stop()

    today = date.today()

    if last_date is None:
        st.warning("Could not detect any date headers in row 1. "
                   "Check that the sheet name and format are correct.")
        st.stop()

    dates_to_fill = _weekdays_between(last_date + timedelta(days=1), today)

    if not dates_to_fill:
        st.success(f"✅ Excel is already up-to-date through **{last_date}**. Nothing to do.")
        st.stop()

    st.subheader("📅 Dates to fill")
    col1, col2 = st.columns(2)
    col1.metric("Last date in Excel", str(last_date))
    col2.metric("Weekdays to fill", len(dates_to_fill))

    with st.expander(f"View all {len(dates_to_fill)} dates"):
        st.write([str(d) for d in dates_to_fill])

    st.markdown("---")

    # ── Action button ─────────────────────────────────────────────────────────
    if not st.button("🚀 Update Excel now", type="primary"):
        st.stop()

    kite_token = kite_token.strip()
    if not kite_token:
        st.error("Not logged in. Click **🔐 Login with Zerodha** in the sidebar (or paste a token manually).")
        st.stop()

    # ── Connect to Kite ───────────────────────────────────────────────────────
    with st.spinner("Connecting to Kite API…"):
        try:
            kite = KiteService(settings)
            kite.client.set_access_token(kite_token)
            # Quick connectivity check
            kite.client.profile()
        except Exception as e:
            st.error(
                f"Kite connection failed: {e}\n\n"
                f"• API key being used: `{api_key_preview}` (from .env) — "
                f"confirm this matches the app on https://developers.kite.trade/apps  \n"
                f"• Access token length pasted: {len(kite_token)} chars "
                f"(a real access_token is ~32 chars; a request_token is also ~32 — make sure "
                f"you used the value returned by `generate_session`, not the one in the redirect URL)"
            )
            st.stop()
    st.success("Connected to Kite ✓")

    # ── Load overrides & workbook ─────────────────────────────────────────────
    overrides = load_overrides(settings.overrides_path)

    wb = load_workbook(workbook_path)

    # ── Progress bar + fill ───────────────────────────────────────────────────
    prog_bar  = st.progress(0.0)
    prog_text = st.empty()

    def progress_cb(frac: float, msg: str):
        prog_bar.progress(frac)
        prog_text.text(msg)

    summary = fill_missing_dates(
        wb=wb,
        sheet_name=sheet_name,
        dates_to_fill=dates_to_fill,
        kite=kite,
        overrides=overrides,
        progress_cb=progress_cb,
    )

    prog_bar.progress(1.0)
    prog_text.text("Saving workbook…")

    # ── Save ──────────────────────────────────────────────────────────────────
    try:
        wb.save(workbook_path)
        wb.close()
    except PermissionError:
        st.error(f"Could not save — please close `{workbook_path.name}` in Excel and try again.")
        st.stop()

    prog_text.text("Done!")

    # ── Results ───────────────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("📋 Results")

    trading_dates  = [d for d in dates_to_fill if summary[d]["status"] == "trading"]
    holiday_dates  = [d for d in dates_to_fill if summary[d]["status"] == "holiday"]
    stocks_failed  = summary.get("stocks_failed", [])

    c1, c2, c3 = st.columns(3)
    c1.metric("Trading days filled", len(trading_dates))
    c2.metric("Holidays / no-data skipped", len(holiday_dates))
    c3.metric("Stocks with errors", len(stocks_failed))

    if trading_dates:
        with st.expander("✅ Trading days updated"):
            for d in trading_dates:
                st.write(f"**{d}** — {summary[d]['updated']} stocks updated, "
                         f"{summary[d]['skipped']} skipped")

    if holiday_dates:
        with st.expander("🚫 Holidays / market-closed days"):
            st.write([str(d) for d in holiday_dates])

    if stocks_failed:
        with st.expander(f"⚠️ {len(stocks_failed)} stocks with errors"):
            for label, reason in stocks_failed:
                st.warning(f"**{label}**: {reason}")

    st.success(f"✅ Workbook saved to `{workbook_path}`")


if __name__ == "__main__":
    main()
