"""
app_deploy.py  –  Deploy on Streamlit Cloud (free tier).
Downloads the Excel from Google Drive, fills missing trading days, uploads back.

Streamlit Secrets required (Settings → Secrets):
    [gcp_service_account]
    type = "service_account"
    project_id = "..."
    private_key_id = "..."
    private_key = "..."
    client_email = "..."
    client_id = "..."
    auth_uri = "..."
    token_uri = "..."
    auth_provider_x509_cert_url = "..."
    client_x509_cert_url = "..."

    [app]
    GDRIVE_FILE_ID  = "your_drive_file_id"
    KITE_API_KEY    = "your_kite_api_key"
    KITE_API_SECRET = "your_kite_api_secret"
    SHEET_NAME      = "Stock Report 2026"
    OVERRIDES_PATH  = "symbol_overrides.json"   # relative to this file

For local testing without Streamlit Cloud, create a credentials.json service-account file
and set GDRIVE_FILE_ID, KITE_API_KEY in your environment.
"""

from __future__ import annotations

import io
import os
import time
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

DATE_COL_MIN_WIDTH = 20.0

# ─────────────────────────────────────────────────────────────────────────────
# Style constants
# ─────────────────────────────────────────────────────────────────────────────
POSITIVE_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
NEGATIVE_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
NEUTRAL_FILL  = PatternFill(fill_type="solid", fgColor="FFEB9C")


# ─────────────────────────────────────────────────────────────────────────────
# Minimal KiteService (no Settings dependency – works without a .env file)
# ─────────────────────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class DailyClose:
    symbol: str
    trading_date: date
    close: float
    open_price: float

    @property
    def change_pct(self) -> float:
        if self.open_price == 0:
            return 0.0
        return (self.close - self.open_price) / self.open_price * 100.0


class KiteClient:
    """Thin wrapper around kiteconnect.KiteConnect."""

    def __init__(self, api_key: str, access_token: str, delay: float = 0.35):
        from kiteconnect import KiteConnect
        self.client = KiteConnect(api_key=api_key)
        self.client.set_access_token(access_token)
        self.delay = delay

    def verify(self):
        return self.client.profile()

    def instruments(self, exchange: str) -> list[dict]:
        return self.client.instruments(exchange)

    def historical_data(self, token: int, from_dt: datetime,
                        to_dt: datetime) -> list[dict]:
        rows = self.client.historical_data(
            instrument_token=token,
            from_date=from_dt,
            to_date=to_dt,
            interval="day",
        )
        time.sleep(self.delay)
        return rows


# ─────────────────────────────────────────────────────────────────────────────
# Google Drive helpers
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_resource(show_spinner=False)
def _get_drive_service():
    """Authenticate and return the Google Drive v3 service (cached)."""
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    scopes = ["https://www.googleapis.com/auth/drive"]

    if "gcp_service_account" in st.secrets:
        info  = dict(st.secrets["gcp_service_account"])
        creds = service_account.Credentials.from_service_account_info(
            info, scopes=scopes
        )
    else:
        # Local fallback
        creds = service_account.Credentials.from_service_account_file(
            "credentials.json", scopes=scopes
        )
    return build("drive", "v3", credentials=creds)


def download_excel_from_drive(file_id: str) -> io.BytesIO:
    """Download a Drive file into a BytesIO buffer."""
    from googleapiclient.http import MediaIoBaseDownload

    service = _get_drive_service()
    request = service.files().get_media(fileId=file_id)
    buf     = io.BytesIO()
    dl      = MediaIoBaseDownload(buf, request)
    done    = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return buf


def upload_excel_to_drive(file_id: str, wb) -> None:
    """Save workbook to BytesIO then overwrite the Drive file."""
    from googleapiclient.http import MediaIoBaseUpload

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    mime    = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    media   = MediaIoBaseUpload(buf, mimetype=mime, resumable=True)
    service = _get_drive_service()
    service.files().update(fileId=file_id, media_body=media).execute()


# ─────────────────────────────────────────────────────────────────────────────
# Date / cell helpers
# ─────────────────────────────────────────────────────────────────────────────

def _parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 30000:
        try:
            return from_excel(value).date()
        except Exception:
            return None
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y",
                    "%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _get_last_date(ws) -> date | None:
    last = None
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            parsed = _parse_date(cell.value)
            if parsed:
                last = parsed
    return last


def _weekdays_between(start: date, end: date) -> list[date]:
    out, cur = [], start
    while cur <= end:
        if cur.weekday() < 5:
            out.append(cur)
        cur += timedelta(days=1)
    return out


def _ensure_min_width(ws, col_idx: int) -> None:
    letter = get_column_letter(col_idx)
    current = ws.column_dimensions[letter].width
    if current is None or current < DATE_COL_MIN_WIDTH:
        ws.column_dimensions[letter].width = DATE_COL_MIN_WIDTH


def _find_or_make_date_col(ws, target: date) -> int:
    last_used = 1
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v not in (None, ""):
            last_used = col
        if _parse_date(v) == target:
            _ensure_min_width(ws, col)
            return col
    new_col = last_used + 1
    hdr  = ws.cell(row=1, column=new_col)
    prev = ws.cell(row=1, column=last_used)
    if prev.has_style:
        hdr.font       = copy(prev.font)
        hdr.fill       = copy(prev.fill)
        hdr.border     = copy(prev.border)
        hdr.alignment  = copy(prev.alignment)
        hdr.protection = copy(prev.protection)
    hdr.value         = target
    hdr.number_format = "dd-mm-yy"
    _ensure_min_width(ws, new_col)
    return new_col


def _write_price(ws, row: int, col: int, close: DailyClose) -> None:
    pct  = close.change_pct
    sign = "+" if pct >= 0 else ""
    cell = ws.cell(row=row, column=col)

    for prev_col in range(col - 1, 1, -1):
        prev = ws.cell(row=row, column=prev_col)
        if prev.value not in (None, ""):
            if prev.has_style:
                cell.font       = copy(prev.font)
                cell.fill       = copy(prev.fill)
                cell.border     = copy(prev.border)
                cell.alignment  = copy(prev.alignment)
                cell.number_format = prev.number_format
                cell.protection = copy(prev.protection)
            break

    cell.value     = f"{close.close:.2f} [{sign}{pct:.2f}%]"
    cell.alignment = Alignment(horizontal="right")

    base_font = cell.font
    if pct > 0:
        cell.fill = copy(POSITIVE_FILL)
        cell.font = Font(color="006100", name=base_font.name, sz=base_font.sz,
                         b=base_font.b, i=base_font.i)
    elif pct < 0:
        cell.fill = copy(NEGATIVE_FILL)
        cell.font = Font(color="9C0006", name=base_font.name, sz=base_font.sz,
                         b=base_font.b, i=base_font.i)
    else:
        cell.fill = copy(NEUTRAL_FILL)
        cell.font = Font(color="9C6500", name=base_font.name, sz=base_font.sz,
                         b=base_font.b, i=base_font.i)


# ─────────────────────────────────────────────────────────────────────────────
# Core fill engine
# ─────────────────────────────────────────────────────────────────────────────

def _build_instrument_map(kite: KiteClient) -> dict[str, int]:
    result = {}
    for exchange in ("NSE", "BSE"):
        for item in kite.instruments(exchange):
            sym = str(item.get("tradingsymbol", "")).upper()
            tok = item.get("instrument_token")
            if sym and tok:
                result.setdefault(sym, int(tok))
    return result


def _resolve_token(sym: str, inst_map: dict[str, int]) -> int | None:
    s = sym.upper()
    return inst_map.get(s) or inst_map.get(f"{s}-EQ") or inst_map.get(f"{s}-BE")


def _load_overrides(path: str) -> dict[str, str]:
    import json
    p = Path(path)
    if not p.exists():
        return {}
    raw = json.loads(p.read_text())
    return {str(k).strip(): str(v).strip().upper() for k, v in raw.items()}


def fill_missing_dates(
    wb,
    sheet_name: str,
    dates_to_fill: list[date],
    kite: KiteClient,
    overrides: dict[str, str],
    progress_cb=None,
) -> dict:
    """
    One API call per stock covering the full date range.
    Holidays are auto-detected: weekdays with no candle data are skipped.
    """
    ws = wb[sheet_name]

    stock_rows: list[tuple[int, str]] = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and val.strip():
            stock_rows.append((row, val.strip()))

    if not dates_to_fill or not stock_rows:
        return {}

    start_dt = dates_to_fill[0]
    end_dt   = dates_to_fill[-1]

    inst_map = _build_instrument_map(kite)

    # ── Pass 1: fetch all data into memory before touching the sheet ─────────
    # Date columns are created only for dates that actually return data,
    # so holidays and "today before market close" do not leave empty columns
    # that the next run would mistakenly treat as already-filled.
    fetched: dict[int, dict[date, DailyClose]] = {row_idx: {} for row_idx, _ in stock_rows}
    stocks_failed: list[tuple[str, str]] = []

    total = len(stock_rows)
    for idx, (row_idx, label) in enumerate(stock_rows):
        if progress_cb:
            progress_cb(idx / total * 0.9, f"Fetching {label} ({idx+1}/{total})…")

        symbol = overrides.get(label, "").upper()
        if not symbol:
            stocks_failed.append((label, "no symbol override defined"))
            continue

        token = _resolve_token(symbol, inst_map)
        if not token:
            stocks_failed.append((label, f"'{symbol}' not found in NSE/BSE instruments"))
            continue

        try:
            records = kite.historical_data(
                token=token,
                from_dt=datetime.combine(start_dt - timedelta(days=1),
                                         datetime.min.time()),
                to_dt=datetime.combine(end_dt, datetime.max.time()),
            )
            by_date: dict[date, dict] = {}
            for rec in records:
                d = rec["date"].date() if isinstance(rec["date"], datetime) else rec["date"]
                by_date[d] = rec

            for target_date in dates_to_fill:
                rec = by_date.get(target_date)
                if rec is not None and float(rec["close"]) > 0:
                    fetched[row_idx][target_date] = DailyClose(
                        symbol=symbol,
                        trading_date=target_date,
                        close=float(rec["close"]),
                        open_price=float(rec["open"]),
                    )

        except Exception as exc:
            stocks_failed.append((label, f"API error: {exc}"))

    # ── Pass 2: create columns only for dates with data, in chronological order ──
    if progress_cb:
        progress_cb(0.95, "Writing to workbook…")

    dates_with_data = sorted({d for row_data in fetched.values() for d in row_data})
    date_to_col: dict[date, int] = {d: _find_or_make_date_col(ws, d) for d in dates_with_data}

    summary: dict = {d: {"status": "holiday", "updated": 0, "skipped": 0}
                     for d in dates_to_fill}
    for d in dates_with_data:
        summary[d]["status"] = "trading"

    for row_idx, row_data in fetched.items():
        for target_date, close_obj in row_data.items():
            _write_price(ws, row_idx, date_to_col[target_date], close_obj)
            summary[target_date]["updated"] += 1

    for d in dates_to_fill:
        summary[d]["skipped"] = len(stock_rows) - summary[d]["updated"]

    summary["stocks_failed"] = stocks_failed
    if progress_cb:
        progress_cb(1.0, "Done.")
    return summary


# ─────────────────────────────────────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────────────────────────────────────

def _get_config() -> dict:
    """Read config from st.secrets (cloud) or environment (local)."""
    try:
        app_cfg = st.secrets.get("app", {})
    except Exception:
        app_cfg = {}

    return {
        "file_id":         app_cfg.get("GDRIVE_FILE_ID",  os.getenv("GDRIVE_FILE_ID", "")),
        "kite_api_key":    app_cfg.get("KITE_API_KEY",    os.getenv("KITE_API_KEY",   "")),
        "kite_api_secret": app_cfg.get("KITE_API_SECRET", os.getenv("KITE_API_SECRET", "")),
        "sheet_name":      app_cfg.get("SHEET_NAME",       os.getenv("STOCK_SHEET_NAME", "Stock Report 2026")),
        "overrides_path":  app_cfg.get("OVERRIDES_PATH",  os.getenv("STOCK_SYMBOL_OVERRIDES",
                                                                      str(Path(__file__).resolve().parent / "symbol_overrides.json"))),
    }


def main():
    st.set_page_config(page_title="Stock Excel Manager – Cloud", page_icon="☁️")
    st.title("☁️ Stock Report Updater – Cloud")
    st.caption("Downloads your Excel from Google Drive, fills missing trading days, uploads back.")

    cfg = _get_config()
    kite_api_key    = cfg["kite_api_key"]
    kite_api_secret = cfg["kite_api_secret"]

    # ── Auto-handle Zerodha OAuth callback (?request_token=… in URL) ──────────
    qp = st.query_params
    if "request_token" in qp and "kite_access_token" not in st.session_state:
        rtok = qp["request_token"]
        if isinstance(rtok, list):
            rtok = rtok[0]
        if not (kite_api_key and kite_api_secret):
            st.error("KITE_API_KEY and KITE_API_SECRET must both be in Streamlit Secrets to auto-exchange.")
            st.query_params.clear()
        else:
            try:
                from kiteconnect import KiteConnect
                kc   = KiteConnect(api_key=kite_api_key)
                data = kc.generate_session(rtok, api_secret=kite_api_secret)
                st.session_state["kite_access_token"] = data["access_token"]
                st.session_state["kite_user_name"]    = data.get("user_name", "")
                st.query_params.clear()
                st.rerun()
            except Exception as e:
                st.error(f"Auto-login failed: {e}")
                st.query_params.clear()

    # ── Sidebar ───────────────────────────────────────────────────────────────
    st.sidebar.header("⚙️ Configuration")

    if "kite_access_token" in st.session_state:
        name = st.session_state.get("kite_user_name", "")
        st.sidebar.success(f"✓ Logged in{(' as ' + name) if name else ''}")
        if st.sidebar.button("🚪 Log out", use_container_width=True):
            st.session_state.pop("kite_access_token", None)
            st.session_state.pop("kite_user_name", None)
            st.rerun()
        kite_token = st.session_state["kite_access_token"]
    else:
        if kite_api_key:
            login_url = (
                f"https://kite.zerodha.com/connect/login?"
                f"api_key={kite_api_key}&v=3"
            )
            st.sidebar.link_button("🔐 Login with Zerodha", login_url, use_container_width=True)
            st.sidebar.caption(
                "Set the Redirect URL on https://developers.kite.trade/apps to this app's "
                "deployed URL (e.g. `https://your-app.streamlit.app/`)."
            )
        else:
            st.sidebar.error("KITE_API_KEY missing in secrets.")
        with st.sidebar.expander("Or paste access token manually"):
            manual = st.text_input("Access Token", type="password", key="manual_tok")
            if st.button("Use this token", key="manual_use") and manual.strip():
                st.session_state["kite_access_token"] = manual.strip()
                st.rerun()
        kite_token = ""

    file_id = cfg["file_id"]
    if not file_id:
        st.error("GDRIVE_FILE_ID is not set. Add it to Streamlit Secrets or environment.")
        st.stop()

    st.info(
        f"**Google Drive File ID:** `{file_id}`  \n"
        f"**Sheet:** `{cfg['sheet_name']}`"
    )

    # ── Preview: download & detect missing dates ───────────────────────────────
    if "wb_buf" not in st.session_state:
        with st.spinner("Downloading Excel from Google Drive…"):
            try:
                st.session_state["wb_buf"] = download_excel_from_drive(file_id)
            except Exception as e:
                st.error(f"Google Drive download failed: {e}")
                st.stop()

    buf = io.BytesIO(st.session_state["wb_buf"].getvalue())
    try:
        wb_preview = load_workbook(buf, read_only=True, data_only=True)
        ws_preview = wb_preview[cfg["sheet_name"]]
        last_date  = _get_last_date(ws_preview)
        wb_preview.close()
    except Exception as e:
        st.error(f"Could not read workbook: {e}")
        st.stop()

    today = date.today()

    if last_date is None:
        st.warning("No date headers found in row 1. Check the sheet name.")
        st.stop()

    dates_to_fill = _weekdays_between(last_date + timedelta(days=1), today)

    if not dates_to_fill:
        st.success(f"✅ Excel is already up-to-date through **{last_date}**. Nothing to do.")
        if st.button("🔄 Refresh from Drive"):
            del st.session_state["wb_buf"]
            st.rerun()
        st.stop()

    st.subheader("📅 Dates to fill")
    c1, c2 = st.columns(2)
    c1.metric("Last date in Excel", str(last_date))
    c2.metric("Weekdays to fill", len(dates_to_fill))

    with st.expander(f"View all {len(dates_to_fill)} dates"):
        st.write([str(d) for d in dates_to_fill])

    st.markdown("---")

    # ── Action button ─────────────────────────────────────────────────────────
    if not st.button("🚀 Update & Upload to Drive", type="primary"):
        st.stop()

    kite_token = kite_token.strip()
    if not kite_token:
        st.error("Not logged in. Click **🔐 Login with Zerodha** in the sidebar (or paste a token manually).")
        st.stop()
    if not kite_api_key:
        st.error("KITE_API_KEY missing in Streamlit secrets.")
        st.stop()

    # ── Connect to Kite ───────────────────────────────────────────────────────
    with st.spinner("Connecting to Kite API…"):
        try:
            kite = KiteClient(api_key=kite_api_key, access_token=kite_token)
            kite.verify()
        except Exception as e:
            api_preview = (kite_api_key[:4] + "…" + kite_api_key[-2:]) \
                if len(kite_api_key) > 6 else kite_api_key
            st.error(
                f"Kite connection failed: {e}\n\n"
                f"• API key being used: `{api_preview}` — confirm this matches the app on "
                f"https://developers.kite.trade/apps  \n"
                f"• Access token length: {len(kite_token)} chars — make sure you pasted the "
                f"value returned by `generate_session`, not the `request_token` from the redirect URL"
            )
            st.stop()
    st.success("Connected to Kite ✓")

    # ── Load workbook from buffer ─────────────────────────────────────────────
    buf = io.BytesIO(st.session_state["wb_buf"].getvalue())
    wb  = load_workbook(buf)

    overrides = _load_overrides(cfg["overrides_path"])
    if not overrides:
        st.error(f"symbol_overrides.json not found at `{cfg['overrides_path']}`")
        st.stop()

    # ── Progress + fill ───────────────────────────────────────────────────────
    prog_bar  = st.progress(0.0)
    prog_text = st.empty()

    def progress_cb(frac: float, msg: str):
        prog_bar.progress(frac)
        prog_text.text(msg)

    summary = fill_missing_dates(
        wb=wb,
        sheet_name=cfg["sheet_name"],
        dates_to_fill=dates_to_fill,
        kite=kite,
        overrides=overrides,
        progress_cb=progress_cb,
    )

    prog_text.text("Uploading to Google Drive…")

    # ── Upload ────────────────────────────────────────────────────────────────
    try:
        upload_excel_to_drive(file_id, wb)
        wb.close()
    except Exception as e:
        st.error(f"Upload failed: {e}")
        st.stop()

    # Clear cached buffer so next run re-downloads the updated file
    del st.session_state["wb_buf"]
    prog_bar.progress(1.0)
    prog_text.text("Uploaded ✓")

    # ── Results ───────────────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("📋 Results")

    trading_dates = [d for d in dates_to_fill if summary[d]["status"] == "trading"]
    holiday_dates = [d for d in dates_to_fill if summary[d]["status"] == "holiday"]
    stocks_failed = summary.get("stocks_failed", [])

    c1, c2, c3 = st.columns(3)
    c1.metric("Trading days filled", len(trading_dates))
    c2.metric("Holidays skipped",    len(holiday_dates))
    c3.metric("Stocks with errors",  len(stocks_failed))

    if trading_dates:
        with st.expander("✅ Trading days updated"):
            for d in trading_dates:
                st.write(f"**{d}** — {summary[d]['updated']} stocks updated, "
                         f"{summary[d]['skipped']} skipped")

    if holiday_dates:
        with st.expander("🚫 Holidays / market-closed days"):
            st.write([str(d) for d in holiday_dates])

    if stocks_failed:
        with st.expander(f"⚠️ {len(stocks_failed)} stocks with errors"):
            for label, reason in stocks_failed:
                st.warning(f"**{label}**: {reason}")

    st.success("✅ Excel updated and uploaded to Google Drive!")


if __name__ == "__main__":
    main()
